import cv2, math, openpyxl
from openpyxl.styles import PatternFill

# ========= Modify below ================
Filepath = "picture-pixel-adjust.png"

# =============== Do not modify ===================
img = cv2.imread(Filepath)
assert img is not None, "file could not be read, check with os.path.exists()"

resize_factor = int(img.shape[0] / 40)
print("Picture size: ", img.shape, ' and type: ', img.dtype)
print("Picture reduce size factor: ", resize_factor)
#print("pixel ::::", img[50,100])

# ------------- Excel
wb = openpyxl.Workbook()
wb.save("picture-to-excel.xlsx")
wb = openpyxl.load_workbook("picture-to-excel.xlsx")
ws = wb.worksheets[0]

for nrow in range(img.shape[0]):
   for ncol in range(img.shape[1]):   
      red, green, blue = img[nrow,ncol]
      if blue > 200 and green > 200 and red > 200: blue, green, red = 0,0,0
      bluehex  = hex(blue)[2:]  if blue  > 15 else '0' + hex(blue)[2:]
      greenhex = hex(green)[2:] if green > 15 else '0' + hex(green)[2:]
      redhex   = hex(red)[2:]   if red   > 15 else '0' + hex(red)[2:]
      hexcolor = '00' + bluehex + greenhex+ redhex
      #print('row: ', excel_row, ' col: ', excel_col, img[nrow,ncol], ' hex: ', hexcolor)
      fillcolor = PatternFill(patternType='solid', fgColor=hexcolor)
      ws.cell(nrow+1,ncol+1).fill = fillcolor
wb.save('picture-to-excel.xlsx')